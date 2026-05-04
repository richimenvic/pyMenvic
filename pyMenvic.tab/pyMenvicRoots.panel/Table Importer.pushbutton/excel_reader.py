# -*- coding: utf-8 -*-

import os
import re
import System
from System.Runtime.InteropServices import Marshal


USED_RANGE_KEY = u"Used Range"
USED_RANGE_DISPLAY = u"Full Worksheet Used Range"


def safe_unicode(value):
    """Return safe Unicode text in IronPython, including Windows-encoded accents."""
    if value is None:
        return u""

    # IronPython 2: unicode is text, str is bytes.
    try:
        if isinstance(value, unicode):
            return value
    except Exception:
        pass

    try:
        if isinstance(value, str):
            for enc in ("utf-8", "cp1252", "latin-1"):
                try:
                    return value.decode(enc)
                except Exception:
                    pass
            try:
                return value.decode("utf-8", "replace")
            except Exception:
                return u""
    except Exception:
        pass

    # .NET strings / COM objects
    try:
        return unicode(value)
    except Exception:
        pass

    try:
        return unicode(value.ToString())
    except Exception:
        pass

    try:
        raw = str(value)
        for enc in ("utf-8", "cp1252", "latin-1"):
            try:
                return raw.decode(enc)
            except Exception:
                pass
    except Exception:
        pass

    return u""


def safe_ascii_text(value):
    """Return plain ASCII-safe UI text to avoid IronPython codepage issues."""
    text = safe_unicode(value)
    if not text:
        return u""
    try:
        import unicodedata
        text = unicodedata.normalize('NFKD', text)
        text = u''.join([c for c in text if not unicodedata.combining(c)])
    except Exception:
        pass
    replacements = {
        u"Ñ": u"N", u"ñ": u"n",
        u"Á": u"A", u"É": u"E", u"Í": u"I", u"Ó": u"O", u"Ú": u"U",
        u"á": u"a", u"é": u"e", u"í": u"i", u"ó": u"o", u"ú": u"u",
    }
    for a, b in replacements.items():
        text = text.replace(a, b)
    cleaned = []
    for ch in text:
        try:
            code = ord(ch)
            if 32 <= code <= 126:
                cleaned.append(ch)
            elif ch in (u"_", u"-", u" "):
                cleaned.append(ch)
        except Exception:
            pass
    return u"".join(cleaned).strip()


def get_last_modified(file_path):
    """Return formatted last modified date."""
    if not file_path or not os.path.exists(file_path):
        return ""
    try:
        timestamp = os.path.getmtime(file_path)
        return System.DateTime.FromFileTimeUtc(
            long(timestamp * 10000000) + 116444736000000000
        ).ToLocalTime().ToString("yyyy-MM-dd HH:mm")
    except Exception:
        return ""


def get_file_name_without_extension(file_path):
    if not file_path:
        return ""
    try:
        return os.path.splitext(os.path.basename(file_path))[0]
    except Exception:
        return ""


def _release_com_object(obj):
    try:
        if obj:
            Marshal.ReleaseComObject(obj)
    except Exception:
        pass


def _get_worksheets_via_com(file_path):
    """Read worksheet names using Excel COM (requires Excel installed)."""
    excel = None
    workbook = None
    try:
        excel_type = System.Type.GetTypeFromProgID("Excel.Application")
        if excel_type is None:
            return None

        excel = System.Activator.CreateInstance(excel_type)
        excel.Visible = False
        excel.DisplayAlerts = False

        workbook = excel.Workbooks.Open(
            file_path,
            False,   # UpdateLinks
            True,    # ReadOnly
        )

        names = []
        for i in range(1, workbook.Worksheets.Count + 1):
            sheet = workbook.Worksheets.Item[i]
            try:
                names.append(sheet.Name)
            finally:
                _release_com_object(sheet)

        return names

    except Exception:
        return None

    finally:
        try:
            if workbook:
                workbook.Close(False)
        except Exception:
            pass
        _release_com_object(workbook)
        try:
            if excel:
                excel.Quit()
        except Exception:
            pass
        _release_com_object(excel)


def _get_worksheets_via_openpyxl(file_path):
    """Read worksheet names using openpyxl (pure Python, no Excel needed)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names
    except Exception:
        return None


def _get_worksheets_via_zipfile(file_path):
    """
    Fallback: read sheet names directly from the xlsx ZIP structure.
    Works for .xlsx without any external library.
    """
    try:
        import zipfile
        import xml.etree.ElementTree as ET

        if not zipfile.is_zipfile(file_path):
            return None

        with zipfile.ZipFile(file_path, 'r') as z:
            if 'xl/workbook.xml' not in z.namelist():
                return None
            with z.open('xl/workbook.xml') as f:
                tree = ET.parse(f)
                root = tree.getroot()

        ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        sheets = root.findall('.//ns:sheet', ns)
        names = [s.get('name') for s in sheets if s.get('name')]
        return names if names else None

    except Exception:
        return None


def get_excel_worksheets(file_path):
    """
    Try multiple methods to read worksheet names.
    1. COM (Excel installed, file not locked)
    2. openpyxl (if installed)
    3. ZIP/XML direct read (always works for .xlsx)
    """
    if not file_path or not os.path.exists(file_path):
        return []

    result = _get_worksheets_via_com(file_path)
    if result is not None:
        return result

    result = _get_worksheets_via_openpyxl(file_path)
    if result is not None:
        return result

    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.xlsx':
        result = _get_worksheets_via_zipfile(file_path)
        if result is not None:
            return result

    return []


def _find_worksheet_by_name(workbook, worksheet_name):
    """COM-safe worksheet lookup. Item[name] is unreliable in IronPython."""
    target = safe_unicode(worksheet_name)
    for i in range(1, workbook.Worksheets.Count + 1):
        sheet = workbook.Worksheets.Item[i]
        try:
            if safe_unicode(sheet.Name) == target:
                return sheet
        except Exception:
            pass
        _release_com_object(sheet)
    return None


def _normalize_excel_address(address):
    if not address:
        return None
    try:
        text = safe_unicode(address).replace("$", "")
        # Excel can return external addresses like '[file.xlsx]Sheet1'!A1:D10.
        if "!" in text:
            text = text.split("!")[-1]
        text = text.replace("'", "")
        return text
    except Exception:
        return None


def _get_used_range_via_com(file_path, worksheet_name):
    excel = None
    workbook = None
    sheet = None
    used_range = None
    try:
        excel_type = System.Type.GetTypeFromProgID("Excel.Application")
        if excel_type is None:
            return None

        excel = System.Activator.CreateInstance(excel_type)
        excel.Visible = False
        excel.DisplayAlerts = False

        workbook = excel.Workbooks.Open(file_path, False, True)
        sheet = _find_worksheet_by_name(workbook, worksheet_name)
        if sheet is None:
            return None

        used_range = sheet.UsedRange

        # Prefer A1-style without dollar signs. Fallback to the default Address.
        try:
            address = used_range.Address(False, False)
        except Exception:
            try:
                address = used_range.Address
            except Exception:
                address = None

        return _normalize_excel_address(address)

    except Exception:
        return None

    finally:
        _release_com_object(used_range)
        _release_com_object(sheet)
        try:
            if workbook:
                workbook.Close(False)
        except Exception:
            pass
        _release_com_object(workbook)
        try:
            if excel:
                excel.Quit()
        except Exception:
            pass
        _release_com_object(excel)


def _get_used_range_via_openpyxl(file_path, worksheet_name):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if worksheet_name not in wb.sheetnames:
            wb.close()
            return None
        ws = wb[worksheet_name]
        dim = ws.calculate_dimension()
        wb.close()
        if dim and dim != "A1:A1":
            return dim
        return dim if dim else None
    except Exception:
        return None


def _column_letter_to_number(col):
    value = 0
    for char in col:
        value = value * 26 + (ord(char.upper()) - ord('A') + 1)
    return value


def _number_to_column_letter(num):
    letters = ""
    while num:
        num, rem = divmod(num - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _range_from_cells(root):
    """Calculate sheet range from cell references if xlsx dimension is missing."""
    min_col = None
    min_row = None
    max_col = None
    max_row = None

    for elem in root.iter():
        ref = elem.get('r')
        if not ref:
            continue
        match = re.match(r'^([A-Z]+)([0-9]+)$', ref)
        if not match:
            continue
        col = _column_letter_to_number(match.group(1))
        row = int(match.group(2))
        min_col = col if min_col is None else min(min_col, col)
        min_row = row if min_row is None else min(min_row, row)
        max_col = col if max_col is None else max(max_col, col)
        max_row = row if max_row is None else max(max_row, row)

    if min_col is None:
        return None

    start_ref = "%s%s" % (_number_to_column_letter(min_col), min_row)
    end_ref = "%s%s" % (_number_to_column_letter(max_col), max_row)
    return start_ref if start_ref == end_ref else "%s:%s" % (start_ref, end_ref)


def _get_sheet_path_from_workbook_relationships(zip_file, sheet_name):
    import posixpath
    import xml.etree.ElementTree as ET

    with zip_file.open('xl/workbook.xml') as f:
        workbook_root = ET.parse(f).getroot()

    main_ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
    rel_ns_uri = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

    rel_id = None
    for sheet in workbook_root.findall('.//ns:sheet', main_ns):
        if sheet.get('name') == sheet_name:
            rel_id = sheet.get('{%s}id' % rel_ns_uri)
            break

    if not rel_id:
        return None

    rels_path = 'xl/_rels/workbook.xml.rels'
    if rels_path not in zip_file.namelist():
        return None

    with zip_file.open(rels_path) as f:
        rels_root = ET.parse(f).getroot()

    for rel in rels_root:
        if rel.get('Id') == rel_id:
            target = rel.get('Target')
            if not target:
                return None
            if target.startswith('/'):
                return target.lstrip('/')
            return posixpath.normpath(posixpath.join('xl', target))

    return None


def _get_used_range_via_zipfile(file_path, worksheet_name):
    """Read sheet dimension from xlsx ZIP without external libs."""
    try:
        import zipfile
        import xml.etree.ElementTree as ET

        if not zipfile.is_zipfile(file_path):
            return None

        with zipfile.ZipFile(file_path, 'r') as z:
            sheet_path = _get_sheet_path_from_workbook_relationships(z, worksheet_name)
            if not sheet_path or sheet_path not in z.namelist():
                return None
            with z.open(sheet_path) as f:
                root = ET.parse(f).getroot()

        ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        dim = root.find('.//ns:dimension', ns)
        if dim is not None:
            ref = dim.get('ref', '')
            if ref:
                return ref

        return _range_from_cells(root)

    except Exception:
        return None


def get_used_range_address(file_path, worksheet_name):
    """
    Try multiple methods to get the used range address (e.g. A1:F25).
    Falls back to 'Used Range' if all methods fail.
    """
    if not file_path or not os.path.exists(file_path):
        return "Used Range"

    result = _get_used_range_via_com(file_path, worksheet_name)
    if result:
        return result

    result = _get_used_range_via_openpyxl(file_path, worksheet_name)
    if result:
        return result

    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.xlsx':
        result = _get_used_range_via_zipfile(file_path, worksheet_name)
        if result:
            return result

    return "Used Range"


def _clean_region_label(label):
    """Return the user-facing region name only, without prefix, address, or Excel internal names."""
    text = safe_unicode(label).strip()
    if not text:
        return u""

    if text == USED_RANGE_KEY or text == USED_RANGE_DISPLAY:
        return USED_RANGE_DISPLAY

    # Remove labels used internally by the reader.
    for prefix in (u"Name ", u"Table "):
        if text.startswith(prefix):
            text = text[len(prefix):].strip()

    # Excel can expose local names as SheetName!RangeName.
    if u"!" in text:
        text = text.split(u"!")[-1].strip()

    text = text.strip(u"'").strip(u'"').strip()

    # Hide Excel internal defined names, e.g. _xlnm._FilterDatabase.
    lower_text = text.lower()
    if lower_text.startswith(u"_xlnm.") or lower_text.startswith(u"_xlnm_") or u"_xlnm." in lower_text:
        return u""

    # Keep the UI/storage codepage-safe. Example: distribución -> distribucion.
    return safe_ascii_text(text)


def _region_unique_key(text):
    """Normalize region labels so Excel Tables and Named Ranges do not appear twice."""
    try:
        value = safe_ascii_text(text).strip().lower()
        # Collapse accidental duplicated spaces/non-breaking spaces from Excel/COM.
        value = value.replace(u"\xa0", u" ")
        while u"  " in value:
            value = value.replace(u"  ", u" ")
        return value
    except Exception:
        return u""


def _safe_add_region(regions, label, address):
    """Add a unique user-facing region name. Address validates only; it is not shown."""
    try:
        # Keep address validation so broken named ranges are ignored,
        # but do not show the address in the UI.
        if address:
            address = _normalize_excel_address(address)
        if not address and safe_unicode(label) != u"Used Range":
            return

        text = _clean_region_label(label)
        if not text:
            return

        new_key = _region_unique_key(text)
        if not new_key:
            return

        for existing in regions:
            if _region_unique_key(existing) == new_key:
                return

        regions.append(text)
    except Exception:
        pass


def _get_regions_via_com(file_path, worksheet_name):
    """Return Used Range + Excel Tables + Named Ranges for a worksheet using COM."""
    excel = None
    workbook = None
    sheet = None
    used_range = None
    regions = []
    try:
        excel_type = System.Type.GetTypeFromProgID("Excel.Application")
        if excel_type is None:
            return None

        excel = System.Activator.CreateInstance(excel_type)
        excel.Visible = False
        excel.DisplayAlerts = False

        workbook = excel.Workbooks.Open(file_path, False, True)
        sheet = _find_worksheet_by_name(workbook, worksheet_name)
        if sheet is None:
            return None

        try:
            used_range = sheet.UsedRange
            try:
                address = used_range.Address(False, False)
            except Exception:
                address = used_range.Address
            _safe_add_region(regions, "Used Range", address)
        except Exception:
            pass

        # Excel Tables / ListObjects on the selected sheet.
        try:
            list_objects = sheet.ListObjects
            for i in range(1, list_objects.Count + 1):
                lo = list_objects.Item(i)
                lo_range = None
                try:
                    lo_range = lo.Range
                    try:
                        address = lo_range.Address(False, False)
                    except Exception:
                        address = lo_range.Address
                    _safe_add_region(regions, u"Table %s" % safe_unicode(lo.Name), address)
                except Exception:
                    pass
                finally:
                    _release_com_object(lo_range)
                    _release_com_object(lo)
            _release_com_object(list_objects)
        except Exception:
            pass

        # Workbook and worksheet scoped named ranges.
        try:
            names = workbook.Names
            for i in range(1, names.Count + 1):
                name_obj = names.Item(i)
                ref_range = None
                try:
                    ref_range = name_obj.RefersToRange
                    try:
                        parent_name = safe_unicode(ref_range.Worksheet.Name)
                    except Exception:
                        parent_name = ""
                    if parent_name == safe_unicode(worksheet_name):
                        try:
                            address = ref_range.Address(False, False)
                        except Exception:
                            address = ref_range.Address
                        name_text = safe_unicode(name_obj.Name)
                        if "!" in name_text:
                            name_text = name_text.split("!")[-1]
                        _safe_add_region(regions, u"Name %s" % safe_unicode(name_text), address)
                except Exception:
                    pass
                finally:
                    _release_com_object(ref_range)
                    _release_com_object(name_obj)
            _release_com_object(names)
        except Exception:
            pass

        return regions if regions else None

    except Exception:
        return None

    finally:
        _release_com_object(used_range)
        _release_com_object(sheet)
        try:
            if workbook:
                workbook.Close(False)
        except Exception:
            pass
        _release_com_object(workbook)
        try:
            if excel:
                excel.Quit()
        except Exception:
            pass
        _release_com_object(excel)


def _get_regions_via_openpyxl(file_path, worksheet_name):
    """Return Used Range + Excel Tables + Named Ranges using openpyxl when available."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
        if worksheet_name not in wb.sheetnames:
            wb.close()
            return None

        ws = wb[worksheet_name]
        regions = []
        dim = ws.calculate_dimension()
        if dim:
            _safe_add_region(regions, "Used Range", dim)

        # Excel tables on this worksheet.
        try:
            table_items = []
            try:
                table_items = ws.tables.items()
            except Exception:
                try:
                    table_items = [(k, ws.tables[k]) for k in ws.tables]
                except Exception:
                    table_items = []

            for name, table in table_items:
                try:
                    ref = table.ref
                except Exception:
                    ref = safe_unicode(table)
                _safe_add_region(regions, u"Table %s" % safe_unicode(name), ref)
        except Exception:
            pass

        # Named ranges. API differs between openpyxl versions, so keep this defensive.
        try:
            defined_names = wb.defined_names
            try:
                iterable = defined_names.definedName
            except Exception:
                try:
                    iterable = defined_names.values()
                except Exception:
                    iterable = []

            for dn in iterable:
                try:
                    name_text = safe_unicode(dn.name)
                    destinations = list(dn.destinations)
                    for title, coord in destinations:
                        if safe_unicode(title) == safe_unicode(worksheet_name):
                            _safe_add_region(regions, u"Name %s" % safe_unicode(name_text), coord)
                except Exception:
                    pass
        except Exception:
            pass

        wb.close()
        return regions if regions else None
    except Exception:
        return None


def _strip_quotes(text):
    try:
        return safe_unicode(text).strip().strip("'").strip('"')
    except Exception:
        return text


def _parse_defined_name_text(text):
    """Parse a definedName value like 'Sheet 1'!$A$1:$D$5."""
    try:
        value = safe_unicode(text).replace("$", "")
        if "!" not in value:
            return None, None
        sheet_part, addr = value.split("!", 1)
        # Ignore formulas, multi-area ranges, and external references for now.
        if "," in addr or "#" in addr or "[" in value:
            return None, None
        return _strip_quotes(sheet_part), _normalize_excel_address(addr)
    except Exception:
        return None, None


def _get_table_paths_for_sheet(zip_file, sheet_path):
    """Return table xml paths referenced by a worksheet in xlsx."""
    try:
        import posixpath
        import xml.etree.ElementTree as ET

        if sheet_path not in zip_file.namelist():
            return []

        with zip_file.open(sheet_path) as f:
            sheet_root = ET.parse(f).getroot()

        rel_ids = []
        rel_ns_uri = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
        for elem in sheet_root.iter():
            if elem.tag.endswith('tablePart'):
                rid = elem.get('{%s}id' % rel_ns_uri)
                if rid:
                    rel_ids.append(rid)

        if not rel_ids:
            return []

        base_dir = posixpath.dirname(sheet_path)
        rels_path = posixpath.join(base_dir, '_rels', posixpath.basename(sheet_path) + '.rels')
        if rels_path not in zip_file.namelist():
            return []

        with zip_file.open(rels_path) as f:
            rels_root = ET.parse(f).getroot()

        paths = []
        for rel in rels_root:
            if rel.get('Id') in rel_ids:
                target = rel.get('Target')
                if not target:
                    continue
                if target.startswith('/'):
                    path = target.lstrip('/')
                else:
                    path = posixpath.normpath(posixpath.join(base_dir, target))
                paths.append(path)
        return paths
    except Exception:
        return []


def _get_regions_via_zipfile(file_path, worksheet_name):
    """Return Used Range + Tables + Named Ranges directly from xlsx XML."""
    try:
        import zipfile
        import xml.etree.ElementTree as ET

        if not zipfile.is_zipfile(file_path):
            return None

        regions = []
        with zipfile.ZipFile(file_path, 'r') as z:
            sheet_path = _get_sheet_path_from_workbook_relationships(z, worksheet_name)
            if sheet_path and sheet_path in z.namelist():
                with z.open(sheet_path) as f:
                    sheet_root = ET.parse(f).getroot()

                ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                dim = sheet_root.find('.//ns:dimension', ns)
                if dim is not None and dim.get('ref'):
                    _safe_add_region(regions, "Used Range", dim.get('ref'))
                else:
                    calc_range = _range_from_cells(sheet_root)
                    if calc_range:
                        _safe_add_region(regions, "Used Range", calc_range)

                # Tables referenced by this worksheet.
                table_paths = _get_table_paths_for_sheet(z, sheet_path)
                for table_path in table_paths:
                    if table_path not in z.namelist():
                        continue
                    try:
                        with z.open(table_path) as f:
                            table_root = ET.parse(f).getroot()
                        table_name = table_root.get('displayName') or table_root.get('name') or os.path.basename(table_path)
                        table_ref = table_root.get('ref')
                        _safe_add_region(regions, u"Table %s" % safe_unicode(table_name), table_ref)
                    except Exception:
                        pass

            # Named ranges from workbook.xml.
            if 'xl/workbook.xml' in z.namelist():
                with z.open('xl/workbook.xml') as f:
                    workbook_root = ET.parse(f).getroot()
                ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                for dn in workbook_root.findall('.//ns:definedName', ns):
                    try:
                        name_text = dn.get('name') or 'Unnamed'
                        sheet_name, addr = _parse_defined_name_text(dn.text)
                        if safe_unicode(sheet_name) == safe_unicode(worksheet_name):
                            _safe_add_region(regions, u"Name %s" % safe_unicode(name_text), addr)
                    except Exception:
                        pass

        return regions if regions else None
    except Exception:
        return None


def get_excel_regions(file_path, worksheet_name):
    """
    Return all usable regions for a worksheet.
    Includes Used Range, Excel Tables/ListObjects, and Named Ranges.
    """
    if not file_path or not os.path.exists(file_path):
        return [USED_RANGE_DISPLAY]

    result = _get_regions_via_com(file_path, worksheet_name)
    if result:
        return result

    result = _get_regions_via_openpyxl(file_path, worksheet_name)
    if result:
        return result

    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.xlsx':
        result = _get_regions_via_zipfile(file_path, worksheet_name)
        if result:
            return result

    used = get_used_range_address(file_path, worksheet_name)
    if used and used != USED_RANGE_KEY:
        return [USED_RANGE_DISPLAY]
    return [USED_RANGE_DISPLAY]

