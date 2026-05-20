# -*- coding: utf-8 -*-

import os
import re
import System
from System.Runtime.InteropServices import Marshal


USED_RANGE_KEY = u"Used Range"
USED_RANGE_DISPLAY = u"Full Worksheet Used Range"
HIDDEN_UNICODE_CHARS = [
    u"\ufeff", u"\u200b", u"\u200c", u"\u200d", u"\u200e", u"\u200f",
    u"\u202a", u"\u202b", u"\u202c", u"\u202d", u"\u202e",
    u"\u2060", u"\u2066", u"\u2067", u"\u2068", u"\u2069",
]


class ExcelTableData(list):
    def __init__(self, rows=None, borders=None, border_available=False, border_method=u"", merged_ranges=None, merges_available=False, column_widths=None, row_heights=None, dimensions_available=False, dimensions_method=u"", dominant_region_font=None, cell_styles=None):
        list.__init__(self, rows or [])
        self.borders = borders
        self.border_available = border_available
        self.border_method = border_method
        self.merged_ranges = merged_ranges or []
        self.merges_available = merges_available
        self.column_widths = column_widths
        self.row_heights = row_heights
        self.dimensions_available = dimensions_available
        self.dimensions_method = dimensions_method
        self.dominant_region_font = dominant_region_font
        self.cell_styles = cell_styles


def clean_hidden_unicode(value):
    try:
        text = value
        for ch in HIDDEN_UNICODE_CHARS:
            text = text.replace(ch, u"")
        cleaned = []
        for ch in text:
            try:
                code = ord(ch)
                if code < 32 and ch not in (u"\t", u"\n", u"\r"):
                    continue
            except Exception:
                pass
            cleaned.append(ch)
        return u"".join(cleaned)
    except Exception:
        return value


def safe_unicode(value):
    """Return safe Unicode text in IronPython, including Windows-encoded accents."""
    if value is None:
        return u""

    # IronPython 2: unicode is text, str is bytes.
    try:
        if isinstance(value, unicode):
            return value
    except Exception:
        pass

    try:
        if isinstance(value, str):
            for enc in ("utf-8", "cp1252", "latin-1"):
                try:
                    return value.decode(enc)
                except Exception:
                    pass
            try:
                return value.decode("utf-8", "replace")
            except Exception:
                return u""
    except Exception:
        pass

    # .NET strings / COM objects
    try:
        return unicode(value)
    except Exception:
        pass

    try:
        return unicode(value.ToString())
    except Exception:
        pass

    try:
        raw = str(value)
        for enc in ("utf-8", "cp1252", "latin-1"):
            try:
                return raw.decode(enc)
            except Exception:
                pass
    except Exception:
        pass

    return u""


def safe_ascii_text(value):
    """Return plain ASCII-safe UI text to avoid IronPython codepage issues."""
    text = safe_unicode(value)
    if not text:
        return u""
    try:
        import unicodedata
        text = unicodedata.normalize('NFKD', text)
        text = u''.join([c for c in text if not unicodedata.combining(c)])
    except Exception:
        pass
    replacements = {
        u"Ñ": u"N", u"ñ": u"n",
        u"Á": u"A", u"É": u"E", u"Í": u"I", u"Ó": u"O", u"Ú": u"U",
        u"á": u"a", u"é": u"e", u"í": u"i", u"ó": u"o", u"ú": u"u",
    }
    for a, b in replacements.items():
        text = text.replace(a, b)
    cleaned = []
    for ch in text:
        try:
            code = ord(ch)
            if 32 <= code <= 126:
                cleaned.append(ch)
            elif ch in (u"_", u"-", u" "):
                cleaned.append(ch)
        except Exception:
            pass
    return u"".join(cleaned).strip()


def get_last_modified(file_path):
    """Return formatted last modified date."""
    if not file_path or not os.path.exists(file_path):
        return ""
    try:
        timestamp = os.path.getmtime(file_path)
        return System.DateTime.FromFileTimeUtc(
            long(timestamp * 10000000) + 116444736000000000
        ).ToLocalTime().ToString("yyyy-MM-dd HH:mm")
    except Exception:
        return ""


def get_file_name_without_extension(file_path):
    if not file_path:
        return ""
    try:
        return os.path.splitext(os.path.basename(file_path))[0]
    except Exception:
        return ""


def _release_com_object(obj):
    try:
        if obj:
            Marshal.ReleaseComObject(obj)
    except Exception:
        pass


def _set_excel_quiet(excel):
    try:
        excel.Visible = False
    except Exception:
        pass
    try:
        excel.DisplayAlerts = False
    except Exception:
        pass


def _get_worksheets_via_com(file_path):
    """Read worksheet names using Excel COM (requires Excel installed)."""
    excel = None
    workbook = None
    try:
        excel_type = System.Type.GetTypeFromProgID("Excel.Application")
        if excel_type is None:
            return None

        excel = System.Activator.CreateInstance(excel_type)
        _set_excel_quiet(excel)

        workbook = excel.Workbooks.Open(
            file_path,
            False,   # UpdateLinks
            True,    # ReadOnly
        )

        names = []
        for i in range(1, workbook.Worksheets.Count + 1):
            sheet = workbook.Worksheets.Item[i]
            try:
                names.append(sheet.Name)
            finally:
                _release_com_object(sheet)

        return names

    except Exception:
        return None

    finally:
        try:
            if workbook:
                workbook.Close(False)
        except Exception:
            pass
        _release_com_object(workbook)
        try:
            if excel:
                excel.Quit()
        except Exception:
            pass
        _release_com_object(excel)


def _get_worksheets_via_openpyxl(file_path):
    """Read worksheet names using openpyxl (pure Python, no Excel needed)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names
    except Exception:
        return None


def _get_worksheets_via_zipfile(file_path):
    """
    Fallback: read sheet names directly from the xlsx ZIP structure.
    Works for .xlsx without any external library.
    """
    try:
        import zipfile
        import xml.etree.ElementTree as ET

        if not zipfile.is_zipfile(file_path):
            return None

        with zipfile.ZipFile(file_path, 'r') as z:
            if 'xl/workbook.xml' not in z.namelist():
                return None
            with z.open('xl/workbook.xml') as f:
                tree = ET.parse(f)
                root = tree.getroot()

        ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        sheets = root.findall('.//ns:sheet', ns)
        names = [s.get('name') for s in sheets if s.get('name')]
        return names if names else None

    except Exception:
        return None


def get_excel_worksheets(file_path):
    """
    Try multiple methods to read worksheet names.
    1. ZIP/XML direct read for .xlsx/.xlsm (no Excel process)\n    2. openpyxl (if installed)\n    3. COM (Excel installed, file not locked)
    """
    if not file_path or not os.path.exists(file_path):
        return []

    ext = os.path.splitext(file_path)[1].lower()
    if ext in ('.xlsx', '.xlsm'):
        result = _get_worksheets_via_zipfile(file_path)
        if result is not None:
            return result

    result = _get_worksheets_via_openpyxl(file_path)
    if result is not None:
        return result

    result = _get_worksheets_via_com(file_path)
    if result is not None:
        return result

    return []


def _find_worksheet_by_name(workbook, worksheet_name):
    """COM-safe worksheet lookup. Item[name] is unreliable in IronPython."""
    target = safe_unicode(worksheet_name)
    for i in range(1, workbook.Worksheets.Count + 1):
        sheet = workbook.Worksheets.Item[i]
        try:
            if safe_unicode(sheet.Name) == target:
                return sheet
        except Exception:
            pass
        _release_com_object(sheet)
    return None


def _normalize_excel_address(address):
    if not address:
        return None
    try:
        text = safe_unicode(address).replace("$", "")
        # Excel can return external addresses like '[file.xlsx]Sheet1'!A1:D10.
        if "!" in text:
            text = text.split("!")[-1]
        text = text.replace("'", "")
        return text
    except Exception:
        return None


def _get_used_range_via_com(file_path, worksheet_name):
    excel = None
    workbook = None
    sheet = None
    used_range = None
    try:
        excel_type = System.Type.GetTypeFromProgID("Excel.Application")
        if excel_type is None:
            return None

        excel = System.Activator.CreateInstance(excel_type)
        _set_excel_quiet(excel)

        workbook = excel.Workbooks.Open(file_path, False, True)
        sheet = _find_worksheet_by_name(workbook, worksheet_name)
        if sheet is None:
            return None

        used_range = sheet.UsedRange

        # Prefer A1-style without dollar signs. Fallback to the default Address.
        try:
            address = used_range.Address(False, False)
        except Exception:
            try:
                address = used_range.Address
            except Exception:
                address = None

        return _normalize_excel_address(address)

    except Exception:
        return None

    finally:
        _release_com_object(used_range)
        _release_com_object(sheet)
        try:
            if workbook:
                workbook.Close(False)
        except Exception:
            pass
        _release_com_object(workbook)
        try:
            if excel:
                excel.Quit()
        except Exception:
            pass
        _release_com_object(excel)


def _get_used_range_via_openpyxl(file_path, worksheet_name):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if worksheet_name not in wb.sheetnames:
            wb.close()
            return None
        ws = wb[worksheet_name]
        dim = ws.calculate_dimension()
        wb.close()
        if dim and dim != "A1:A1":
            return dim
        return dim if dim else None
    except Exception:
        return None


def _column_letter_to_number(col):
    value = 0
    for char in col:
        value = value * 26 + (ord(char.upper()) - ord('A') + 1)
    return value


def _number_to_column_letter(num):
    letters = ""
    while num:
        num, rem = divmod(num - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _range_from_cells(root):
    """Calculate sheet range from cell references if xlsx dimension is missing."""
    min_col = None
    min_row = None
    max_col = None
    max_row = None

    for elem in root.iter():
        ref = elem.get('r')
        if not ref:
            continue
        match = re.match(r'^([A-Z]+)([0-9]+)$', ref)
        if not match:
            continue
        col = _column_letter_to_number(match.group(1))
        row = int(match.group(2))
        min_col = col if min_col is None else min(min_col, col)
        min_row = row if min_row is None else min(min_row, row)
        max_col = col if max_col is None else max(max_col, col)
        max_row = row if max_row is None else max(max_row, row)

    if min_col is None:
        return None

    start_ref = "%s%s" % (_number_to_column_letter(min_col), min_row)
    end_ref = "%s%s" % (_number_to_column_letter(max_col), max_row)
    return start_ref if start_ref == end_ref else "%s:%s" % (start_ref, end_ref)


def _get_sheet_path_from_workbook_relationships(zip_file, sheet_name):
    import posixpath
    import xml.etree.ElementTree as ET

    with zip_file.open('xl/workbook.xml') as f:
        workbook_root = ET.parse(f).getroot()

    main_ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
    rel_ns_uri = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

    rel_id = None
    for sheet in workbook_root.findall('.//ns:sheet', main_ns):
        if sheet.get('name') == sheet_name:
            rel_id = sheet.get('{%s}id' % rel_ns_uri)
            break

    if not rel_id:
        return None

    rels_path = 'xl/_rels/workbook.xml.rels'
    if rels_path not in zip_file.namelist():
        return None

    with zip_file.open(rels_path) as f:
        rels_root = ET.parse(f).getroot()

    for rel in rels_root:
        if rel.get('Id') == rel_id:
            target = rel.get('Target')
            if not target:
                return None
            if target.startswith('/'):
                return target.lstrip('/')
            return posixpath.normpath(posixpath.join('xl', target))

    return None


def _get_used_range_via_zipfile(file_path, worksheet_name):
    """Read sheet dimension from xlsx ZIP without external libs."""
    try:
        import zipfile
        import xml.etree.ElementTree as ET

        if not zipfile.is_zipfile(file_path):
            return None

        with zipfile.ZipFile(file_path, 'r') as z:
            sheet_path = _get_sheet_path_from_workbook_relationships(z, worksheet_name)
            if not sheet_path or sheet_path not in z.namelist():
                return None
            with z.open(sheet_path) as f:
                root = ET.parse(f).getroot()

        ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        dim = root.find('.//ns:dimension', ns)
        if dim is not None:
            ref = dim.get('ref', '')
            if ref:
                return ref

        return _range_from_cells(root)

    except Exception:
        return None


def get_used_range_address(file_path, worksheet_name):
    """
    Try multiple methods to get the used range address (e.g. A1:F25).
    Falls back to 'Used Range' if all methods fail.
    """
    if not file_path or not os.path.exists(file_path):
        return "Used Range"

    ext = os.path.splitext(file_path)[1].lower()
    if ext in ('.xlsx', '.xlsm'):
        result = _get_used_range_via_zipfile(file_path, worksheet_name)
        if result:
            return result

    result = _get_used_range_via_openpyxl(file_path, worksheet_name)
    if result:
        return result

    result = _get_used_range_via_com(file_path, worksheet_name)
    if result:
        return result

    return "Used Range"


def _clean_region_label(label):
    """Return the user-facing region name only, without prefix, address, or Excel internal names."""
    text = safe_unicode(label).strip()
    if not text:
        return u""

    if text == USED_RANGE_KEY or text == USED_RANGE_DISPLAY:
        return USED_RANGE_DISPLAY

    # Remove labels used internally by the reader.
    for prefix in (u"Name ", u"Table "):
        if text.startswith(prefix):
            text = text[len(prefix):].strip()

    # Excel can expose local names as SheetName!RangeName.
    if u"!" in text:
        text = text.split(u"!")[-1].strip()

    text = text.strip(u"'").strip(u'"').strip()

    # Hide Excel internal defined names, e.g. _xlnm._FilterDatabase.
    lower_text = text.lower()
    if lower_text.startswith(u"_xlnm.") or lower_text.startswith(u"_xlnm_") or u"_xlnm." in lower_text:
        return u""

    # Keep the UI/storage codepage-safe. Example: distribución -> distribucion.
    return safe_ascii_text(text)


def _region_unique_key(text):
    """Normalize region labels so Excel Tables and Named Ranges do not appear twice."""
    try:
        value = safe_ascii_text(text).strip().lower()
        # Collapse accidental duplicated spaces/non-breaking spaces from Excel/COM.
        value = value.replace(u"\xa0", u" ")
        while u"  " in value:
            value = value.replace(u"  ", u" ")
        return value
    except Exception:
        return u""


def _safe_add_region(regions, label, address):
    """Add a unique user-facing region name. Address validates only; it is not shown."""
    try:
        # Keep address validation so broken named ranges are ignored,
        # but do not show the address in the UI.
        if address:
            address = _normalize_excel_address(address)
        if not address and safe_unicode(label) != u"Used Range":
            return

        text = _clean_region_label(label)
        if not text:
            return

        new_key = _region_unique_key(text)
        if not new_key:
            return

        for existing in regions:
            if _region_unique_key(existing) == new_key:
                return

        regions.append(text)
    except Exception:
        pass


def _is_used_range_region(label):
    try:
        text = _clean_region_label(label)
        return _region_unique_key(text) == _region_unique_key(USED_RANGE_DISPLAY)
    except Exception:
        return False


def _filter_region_list_for_ui(regions):
    useful = []
    for region in regions or []:
        try:
            if _is_used_range_region(region):
                continue
            text = _clean_region_label(region)
            if text:
                useful.append(text)
        except Exception:
            pass
    if useful:
        return useful
    return [USED_RANGE_DISPLAY]


def _get_regions_via_com(file_path, worksheet_name):
    """Return Used Range + Excel Tables + Named Ranges for a worksheet using COM."""
    excel = None
    workbook = None
    sheet = None
    used_range = None
    regions = []
    try:
        excel_type = System.Type.GetTypeFromProgID("Excel.Application")
        if excel_type is None:
            return None

        excel = System.Activator.CreateInstance(excel_type)
        _set_excel_quiet(excel)

        workbook = excel.Workbooks.Open(file_path, False, True)
        sheet = _find_worksheet_by_name(workbook, worksheet_name)
        if sheet is None:
            return None

        try:
            used_range = sheet.UsedRange
            try:
                address = used_range.Address(False, False)
            except Exception:
                address = used_range.Address
            _safe_add_region(regions, "Used Range", address)
        except Exception:
            pass

        # Excel Tables / ListObjects on the selected sheet.
        try:
            list_objects = sheet.ListObjects
            for i in range(1, list_objects.Count + 1):
                lo = list_objects.Item(i)
                lo_range = None
                try:
                    lo_range = lo.Range
                    try:
                        address = lo_range.Address(False, False)
                    except Exception:
                        address = lo_range.Address
                    _safe_add_region(regions, u"Table %s" % safe_unicode(lo.Name), address)
                except Exception:
                    pass
                finally:
                    _release_com_object(lo_range)
                    _release_com_object(lo)
            _release_com_object(list_objects)
        except Exception:
            pass

        # Workbook and worksheet scoped named ranges.
        try:
            names = workbook.Names
            for i in range(1, names.Count + 1):
                name_obj = names.Item(i)
                ref_range = None
                try:
                    ref_range = name_obj.RefersToRange
                    try:
                        parent_name = safe_unicode(ref_range.Worksheet.Name)
                    except Exception:
                        parent_name = ""
                    if parent_name == safe_unicode(worksheet_name):
                        try:
                            address = ref_range.Address(False, False)
                        except Exception:
                            address = ref_range.Address
                        name_text = safe_unicode(name_obj.Name)
                        if "!" in name_text:
                            name_text = name_text.split("!")[-1]
                        if name_text.lower() == u"_xlnm.print_area":
                            _safe_add_region(regions, u"Print Area", address)
                        else:
                            _safe_add_region(regions, u"Name %s" % safe_unicode(name_text), address)
                except Exception:
                    pass
                finally:
                    _release_com_object(ref_range)
                    _release_com_object(name_obj)
            _release_com_object(names)
        except Exception:
            pass

        return regions if regions else None

    except Exception:
        return None

    finally:
        _release_com_object(used_range)
        _release_com_object(sheet)
        try:
            if workbook:
                workbook.Close(False)
        except Exception:
            pass
        _release_com_object(workbook)
        try:
            if excel:
                excel.Quit()
        except Exception:
            pass
        _release_com_object(excel)


def _get_regions_via_openpyxl(file_path, worksheet_name):
    """Return Used Range + Excel Tables + Named Ranges using openpyxl when available."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
        if worksheet_name not in wb.sheetnames:
            wb.close()
            return None

        ws = wb[worksheet_name]
        regions = []
        dim = ws.calculate_dimension()
        if dim:
            _safe_add_region(regions, "Used Range", dim)

        # Excel tables on this worksheet.
        try:
            table_items = []
            try:
                table_items = ws.tables.items()
            except Exception:
                try:
                    table_items = [(k, ws.tables[k]) for k in ws.tables]
                except Exception:
                    table_items = []

            for name, table in table_items:
                try:
                    ref = table.ref
                except Exception:
                    ref = safe_unicode(table)
                _safe_add_region(regions, u"Table %s" % safe_unicode(name), ref)
        except Exception:
            pass

        # Named ranges. API differs between openpyxl versions, so keep this defensive.
        try:
            defined_names = wb.defined_names
            try:
                iterable = defined_names.definedName
            except Exception:
                try:
                    iterable = defined_names.values()
                except Exception:
                    iterable = []

            for dn in iterable:
                try:
                    name_text = safe_unicode(dn.name)
                    destinations = list(dn.destinations)
                    for title, coord in destinations:
                        if safe_unicode(title) == safe_unicode(worksheet_name):
                            if name_text.lower() == u"_xlnm.print_area":
                                _safe_add_region(regions, u"Print Area", coord)
                            else:
                                _safe_add_region(regions, u"Name %s" % safe_unicode(name_text), coord)
                except Exception:
                    pass
        except Exception:
            pass

        wb.close()
        return regions if regions else None
    except Exception:
        return None


def _strip_quotes(text):
    try:
        return safe_unicode(text).strip().strip("'").strip('"')
    except Exception:
        return text


def _parse_defined_name_text(text):
    """Parse a definedName value like 'Sheet 1'!$A$1:$D$5."""
    try:
        value = safe_unicode(text).replace("$", "")
        if "!" not in value:
            return None, None
        sheet_part, addr = value.split("!", 1)
        # Ignore formulas, multi-area ranges, and external references for now.
        if "," in addr or "#" in addr or "[" in value:
            return None, None
        return _strip_quotes(sheet_part), _normalize_excel_address(addr)
    except Exception:
        return None, None


def _get_table_paths_for_sheet(zip_file, sheet_path):
    """Return table xml paths referenced by a worksheet in xlsx."""
    try:
        import posixpath
        import xml.etree.ElementTree as ET

        if sheet_path not in zip_file.namelist():
            return []

        with zip_file.open(sheet_path) as f:
            sheet_root = ET.parse(f).getroot()

        rel_ids = []
        rel_ns_uri = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
        for elem in sheet_root.iter():
            if elem.tag.endswith('tablePart'):
                rid = elem.get('{%s}id' % rel_ns_uri)
                if rid:
                    rel_ids.append(rid)

        if not rel_ids:
            return []

        base_dir = posixpath.dirname(sheet_path)
        rels_path = posixpath.join(base_dir, '_rels', posixpath.basename(sheet_path) + '.rels')
        if rels_path not in zip_file.namelist():
            return []

        with zip_file.open(rels_path) as f:
            rels_root = ET.parse(f).getroot()

        paths = []
        for rel in rels_root:
            if rel.get('Id') in rel_ids:
                target = rel.get('Target')
                if not target:
                    continue
                if target.startswith('/'):
                    path = target.lstrip('/')
                else:
                    path = posixpath.normpath(posixpath.join(base_dir, target))
                paths.append(path)
        return paths
    except Exception:
        return []


def _get_regions_via_zipfile(file_path, worksheet_name):
    """Return Used Range + Tables + Named Ranges directly from xlsx XML."""
    try:
        import zipfile
        import xml.etree.ElementTree as ET

        if not zipfile.is_zipfile(file_path):
            return None

        regions = []
        with zipfile.ZipFile(file_path, 'r') as z:
            sheet_path = _get_sheet_path_from_workbook_relationships(z, worksheet_name)
            if sheet_path and sheet_path in z.namelist():
                with z.open(sheet_path) as f:
                    sheet_root = ET.parse(f).getroot()

                ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                dim = sheet_root.find('.//ns:dimension', ns)
                if dim is not None and dim.get('ref'):
                    _safe_add_region(regions, "Used Range", dim.get('ref'))
                else:
                    calc_range = _range_from_cells(sheet_root)
                    if calc_range:
                        _safe_add_region(regions, "Used Range", calc_range)

                # Tables referenced by this worksheet.
                table_paths = _get_table_paths_for_sheet(z, sheet_path)
                for table_path in table_paths:
                    if table_path not in z.namelist():
                        continue
                    try:
                        with z.open(table_path) as f:
                            table_root = ET.parse(f).getroot()
                        table_name = table_root.get('displayName') or table_root.get('name') or os.path.basename(table_path)
                        table_ref = table_root.get('ref')
                        _safe_add_region(regions, u"Table %s" % safe_unicode(table_name), table_ref)
                    except Exception:
                        pass

            # Named ranges from workbook.xml.
            if 'xl/workbook.xml' in z.namelist():
                with z.open('xl/workbook.xml') as f:
                    workbook_root = ET.parse(f).getroot()
                ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                for dn in workbook_root.findall('.//ns:definedName', ns):
                    try:
                        name_text = dn.get('name') or 'Unnamed'
                        sheet_name, addr = _parse_defined_name_text(dn.text)
                        if safe_unicode(sheet_name) == safe_unicode(worksheet_name):
                            if safe_unicode(name_text).lower() == u"_xlnm.print_area":
                                _safe_add_region(regions, u"Print Area", addr)
                            else:
                                _safe_add_region(regions, u"Name %s" % safe_unicode(name_text), addr)
                    except Exception:
                        pass

        return regions if regions else None
    except Exception:
        return None


def get_excel_regions(file_path, worksheet_name):
    """
    Return all usable regions for a worksheet.
    Includes Used Range, Excel Tables/ListObjects, and Named Ranges.
    """
    if not file_path or not os.path.exists(file_path):
        return [USED_RANGE_DISPLAY]

    ext = os.path.splitext(file_path)[1].lower()
    if ext in ('.xlsx', '.xlsm'):
        result = _get_regions_via_zipfile(file_path, worksheet_name)
        if result:
            return _filter_region_list_for_ui(result)

    result = _get_regions_via_openpyxl(file_path, worksheet_name)
    if result:
        return _filter_region_list_for_ui(result)

    result = _get_regions_via_com(file_path, worksheet_name)
    if result:
        return _filter_region_list_for_ui(result)

    used = get_used_range_address(file_path, worksheet_name)
    if used and used != USED_RANGE_KEY:
        return [USED_RANGE_DISPLAY]
    return [USED_RANGE_DISPLAY]


def _clean_cell_value(value):
    if value is None:
        return u""
    try:
        return safe_unicode(value).strip()
    except Exception:
        return u""


def _normalize_region_key(label):
    return _region_unique_key(_clean_region_label(clean_hidden_unicode(safe_unicode(label))))


def _looks_like_excel_address(text):
    try:
        value = safe_unicode(text).replace("$", "").strip()
        if "!" in value:
            value = value.split("!")[-1]
        value = value.strip("'").strip('"')
        return re.match(r'^[A-Z]{1,3}[0-9]+(:[A-Z]{1,3}[0-9]+)?$', value.upper()) is not None
    except Exception:
        return False


def _extract_legacy_region_address(region_display):
    try:
        text = safe_unicode(region_display).replace("$", "")
        if ":" not in text:
            return None
        match = re.search(r'([A-Z]{1,3}[0-9]+:[A-Z]{1,3}[0-9]+)', text.upper())
        if match:
            return _normalize_excel_address(match.group(1))
        match = re.search(r'([A-Z]{1,3}[0-9]+)', text.upper())
        if match and _looks_like_excel_address(match.group(1)):
            return _normalize_excel_address(match.group(1))
    except Exception:
        pass
    return None


def _find_address_by_clean_key(addresses, region_display):
    try:
        if not addresses:
            return None
        target_key = _normalize_region_key(region_display)
        if target_key in addresses:
            return addresses[target_key]

        # Be lenient with stored/display strings from older versions.
        target_raw = _region_unique_key(region_display)
        for key, address in addresses.items():
            if key == target_raw:
                return address
            if _normalize_region_key(key) == target_key:
                return address
    except Exception:
        pass
    return None


def _add_region_address(addresses, label, address):
    try:
        clean_label = _clean_region_label(label)
        key = _normalize_region_key(clean_label)
        address = _normalize_excel_address(address)
        if key and address:
            addresses[key] = address
    except Exception:
        pass


def _get_region_addresses_via_openpyxl(file_path, worksheet_name):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
        if worksheet_name not in wb.sheetnames:
            wb.close()
            return None

        ws = wb[worksheet_name]
        addresses = {}
        _add_region_address(addresses, USED_RANGE_DISPLAY, ws.calculate_dimension())

        try:
            table_items = []
            try:
                table_items = ws.tables.items()
            except Exception:
                try:
                    table_items = [(k, ws.tables[k]) for k in ws.tables]
                except Exception:
                    table_items = []
            for name, table in table_items:
                try:
                    ref = table.ref
                except Exception:
                    ref = safe_unicode(table)
                _add_region_address(addresses, u"Table %s" % safe_unicode(name), ref)
        except Exception:
            pass

        try:
            defined_names = wb.defined_names
            try:
                iterable = defined_names.definedName
            except Exception:
                try:
                    iterable = defined_names.values()
                except Exception:
                    iterable = []
            for dn in iterable:
                try:
                    name_text = safe_unicode(dn.name)
                    destinations = list(dn.destinations)
                    for title, coord in destinations:
                        if safe_unicode(title) == safe_unicode(worksheet_name):
                            if name_text.lower() == u"_xlnm.print_area":
                                _add_region_address(addresses, u"Print Area", coord)
                            else:
                                _add_region_address(addresses, u"Name %s" % name_text, coord)
                except Exception:
                    pass
        except Exception:
            pass

        wb.close()
        return addresses
    except Exception:
        return None


def _get_region_addresses_via_zipfile(file_path, worksheet_name):
    try:
        import zipfile
        import xml.etree.ElementTree as ET

        if not zipfile.is_zipfile(file_path):
            return None

        addresses = {}
        with zipfile.ZipFile(file_path, 'r') as z:
            sheet_path = _get_sheet_path_from_workbook_relationships(z, worksheet_name)
            if sheet_path and sheet_path in z.namelist():
                with z.open(sheet_path) as f:
                    sheet_root = ET.parse(f).getroot()
                ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                dim = sheet_root.find('.//ns:dimension', ns)
                if dim is not None and dim.get('ref'):
                    _add_region_address(addresses, USED_RANGE_DISPLAY, dim.get('ref'))
                else:
                    _add_region_address(addresses, USED_RANGE_DISPLAY, _range_from_cells(sheet_root))

                for table_path in _get_table_paths_for_sheet(z, sheet_path):
                    if table_path not in z.namelist():
                        continue
                    try:
                        with z.open(table_path) as f:
                            table_root = ET.parse(f).getroot()
                        table_name = table_root.get('displayName') or table_root.get('name') or os.path.basename(table_path)
                        _add_region_address(addresses, u"Table %s" % safe_unicode(table_name), table_root.get('ref'))
                    except Exception:
                        pass

            if 'xl/workbook.xml' in z.namelist():
                with z.open('xl/workbook.xml') as f:
                    workbook_root = ET.parse(f).getroot()
                ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                for dn in workbook_root.findall('.//ns:definedName', ns):
                    try:
                        name_text = dn.get('name') or 'Unnamed'
                        sheet_name, addr = _parse_defined_name_text(dn.text)
                        if safe_unicode(sheet_name) == safe_unicode(worksheet_name):
                            if safe_unicode(name_text).lower() == u"_xlnm.print_area":
                                _add_region_address(addresses, u"Print Area", addr)
                            else:
                                _add_region_address(addresses, u"Name %s" % safe_unicode(name_text), addr)
                    except Exception:
                        pass
        return addresses
    except Exception:
        return None


def _get_region_address_via_com(file_path, worksheet_name, region):
    excel = None
    workbook = None
    sheet = None
    try:
        excel_type = System.Type.GetTypeFromProgID("Excel.Application")
        if excel_type is None:
            return None
        excel = System.Activator.CreateInstance(excel_type)
        _set_excel_quiet(excel)
        workbook = excel.Workbooks.Open(file_path, False, True)
        sheet = _find_worksheet_by_name(workbook, worksheet_name)
        if sheet is None:
            return None

        target_key = _normalize_region_key(region)
        if not target_key or target_key == _normalize_region_key(USED_RANGE_DISPLAY):
            try:
                return _normalize_excel_address(sheet.UsedRange.Address(False, False))
            except Exception:
                return None

        try:
            list_objects = sheet.ListObjects
            for i in range(1, list_objects.Count + 1):
                lo = list_objects.Item(i)
                lo_range = None
                try:
                    if _normalize_region_key(lo.Name) == target_key:
                        lo_range = lo.Range
                        return _normalize_excel_address(lo_range.Address(False, False))
                except Exception:
                    pass
                finally:
                    _release_com_object(lo_range)
                    _release_com_object(lo)
            _release_com_object(list_objects)
        except Exception:
            pass

        try:
            names = workbook.Names
            for i in range(1, names.Count + 1):
                name_obj = names.Item(i)
                ref_range = None
                try:
                    ref_range = name_obj.RefersToRange
                    if safe_unicode(ref_range.Worksheet.Name) == safe_unicode(worksheet_name):
                        name_text = safe_unicode(name_obj.Name)
                        if "!" in name_text:
                            name_text = name_text.split("!")[-1]
                        label = u"Print Area" if name_text.lower() == u"_xlnm.print_area" else name_text
                        if _normalize_region_key(label) == target_key:
                            return _normalize_excel_address(ref_range.Address(False, False))
                except Exception:
                    pass
                finally:
                    _release_com_object(ref_range)
                    _release_com_object(name_obj)
            _release_com_object(names)
        except Exception:
            pass
    except Exception:
        return None
    finally:
        _release_com_object(sheet)
        try:
            if workbook:
                workbook.Close(False)
        except Exception:
            pass
        _release_com_object(workbook)
        try:
            if excel:
                excel.Quit()
        except Exception:
            pass
        _release_com_object(excel)
    return None


def _resolve_region_address_with_method(file_path, worksheet_name, region_display):
    file_path = safe_unicode(file_path)
    worksheet_name = safe_unicode(worksheet_name)
    region_display = safe_unicode(region_display) or USED_RANGE_DISPLAY

    legacy_address = _extract_legacy_region_address(region_display)
    if legacy_address:
        return legacy_address, "legacy display string", ""

    if _looks_like_excel_address(region_display):
        return _normalize_excel_address(region_display), "direct address", ""

    key = _normalize_region_key(region_display)
    used_key = _normalize_region_key(USED_RANGE_DISPLAY)
    if not key or key == used_key or safe_unicode(region_display) == USED_RANGE_KEY:
        address = get_used_range_address(file_path, worksheet_name)
        if address and address != USED_RANGE_KEY:
            return address, "used range", ""
        return None, "used range", "Could not resolve worksheet used range."

    ext = os.path.splitext(file_path)[1].lower()
    if ext in ('.xlsx', '.xlsm'):
        addresses = _get_region_addresses_via_zipfile(file_path, worksheet_name)
        result = _find_address_by_clean_key(addresses, region_display)
        if result:
            return result, "xlsx xml", ""

    addresses = _get_region_addresses_via_openpyxl(file_path, worksheet_name)
    result = _find_address_by_clean_key(addresses, region_display)
    if result:
        return result, "openpyxl", ""

    result = _get_region_address_via_com(file_path, worksheet_name, region_display)
    if result:
        return result, "Excel COM", ""

    return None, "not resolved", "No matching named range, Excel table, or print area was found."


def resolve_region_to_address(file_path, worksheet_name, region_display):
    address, method, reason = _resolve_region_address_with_method(file_path, worksheet_name, region_display)
    return address


def _address_to_bounds(address):
    try:
        text = _normalize_excel_address(address)
        if not text:
            return None
        parts = text.split(":")
        if len(parts) == 1:
            parts.append(parts[0])
        match1 = re.match(r'^([A-Z]+)([0-9]+)$', parts[0].upper())
        match2 = re.match(r'^([A-Z]+)([0-9]+)$', parts[1].upper())
        if not match1 or not match2:
            return None
        min_col = _column_letter_to_number(match1.group(1))
        min_row = int(match1.group(2))
        max_col = _column_letter_to_number(match2.group(1))
        max_row = int(match2.group(2))
        if max_col < min_col:
            min_col, max_col = max_col, min_col
        if max_row < min_row:
            min_row, max_row = max_row, min_row
        return min_col, min_row, max_col, max_row
    except Exception:
        return None


def _read_cell_values_via_openpyxl(file_path, worksheet_name, address):
    wb = None
    try:
        import openpyxl
        try:
            from openpyxl.utils.cell import range_boundaries
        except Exception:
            range_boundaries = None

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if worksheet_name not in wb.sheetnames:
            wb.close()
            return None, "Worksheet was not found by openpyxl."
        ws = wb[worksheet_name]

        if range_boundaries:
            min_col, min_row, max_col, max_row = range_boundaries(address)
        else:
            bounds = _address_to_bounds(address)
            if not bounds:
                wb.close()
                return None, "Openpyxl could not parse address '%s'." % safe_unicode(address)
            min_col, min_row, max_col, max_row = bounds

        data = []
        cell_styles = []
        for row_index in range(min_row, max_row + 1):
            row_values = []
            row_styles = []
            for col_index in range(min_col, max_col + 1):
                cell = ws.cell(row=row_index, column=col_index)
                row_values.append(_clean_cell_value(cell.value))
                row_styles.append(_cell_style_dict(cell))
            data.append(row_values)
            cell_styles.append(row_styles)
        column_widths, row_heights = _read_openpyxl_dimensions(ws, min_col, min_row, max_col, max_row)
        wb.close()
        return ExcelTableData(data, None, False, u"openpyxl values", None, False, column_widths, row_heights, column_widths is not None or row_heights is not None, u"openpyxl", None, cell_styles if _has_useful_cell_styles(cell_styles) else None), ""
    except Exception as ex:
        try:
            if wb:
                wb.close()
        except Exception:
            pass
        return None, safe_unicode(ex)


def _cell_side_has_border(side):
    try:
        if side is None:
            return False
        style = getattr(side, "style", None)
        if style:
            return True
    except Exception:
        pass
    return False


def _cell_border_dict(cell):
    result = {"top": False, "right": False, "bottom": False, "left": False}
    try:
        border = cell.border
        result["top"] = _cell_side_has_border(border.top)
        result["right"] = _cell_side_has_border(border.right)
        result["bottom"] = _cell_side_has_border(border.bottom)
        result["left"] = _cell_side_has_border(border.left)
    except Exception:
        pass
    return result


def _add_font_count(font_counts, font_name):
    try:
        name = clean_hidden_unicode(safe_unicode(font_name)).strip()
    except Exception:
        name = u""
    if not name:
        return
    try:
        font_counts[name] = font_counts.get(name, 0) + 1
    except Exception:
        pass


def _dominant_font_from_counts(font_counts):
    dominant_name = None
    dominant_count = 0
    try:
        for name, count in font_counts.items():
            if count > dominant_count:
                dominant_name = name
                dominant_count = count
    except Exception:
        pass
    return dominant_name


def _cell_style_dict(cell):
    """Return a small, Revit-friendly style description for one Excel cell."""
    result = {
        "font_name": u"",
        "size_pt": None,
        "bold": False,
        "italic": False,
        "horizontal": u"",
        "vertical": u"",
    }
    try:
        font = cell.font
        try:
            result["font_name"] = clean_hidden_unicode(safe_unicode(font.name)).strip()
        except Exception:
            result["font_name"] = u""
        try:
            if font.sz is not None:
                result["size_pt"] = float(font.sz)
        except Exception:
            result["size_pt"] = None
        try:
            result["bold"] = bool(font.bold)
        except Exception:
            result["bold"] = False
        try:
            result["italic"] = bool(font.italic)
        except Exception:
            result["italic"] = False
    except Exception:
        pass
    try:
        alignment = cell.alignment
        try:
            result["horizontal"] = clean_hidden_unicode(safe_unicode(alignment.horizontal)).strip()
        except Exception:
            result["horizontal"] = u""
        try:
            result["vertical"] = clean_hidden_unicode(safe_unicode(alignment.vertical)).strip()
        except Exception:
            result["vertical"] = u""
    except Exception:
        pass
    return result

def _has_useful_cell_styles(cell_styles):
    try:
        for row in cell_styles or []:
            for style in row or []:
                if not style:
                    continue
                if style.get("font_name") or style.get("size_pt") or style.get("bold") or style.get("italic") or style.get("horizontal") or style.get("vertical"):
                    return True
    except Exception:
        pass
    return False


def _read_openpyxl_dimensions(ws, min_col, min_row, max_col, max_row):
    column_widths = []
    row_heights = []
    try:
        default_col_width = None
        try:
            default_col_width = ws.sheet_format.defaultColWidth
        except Exception:
            default_col_width = None
        for col_index in range(min_col, max_col + 1):
            width = None
            try:
                col_letter = _number_to_column_letter(col_index)
                width = ws.column_dimensions[col_letter].width
            except Exception:
                width = None
            if width is None:
                width = default_col_width
            try:
                column_widths.append(float(width) if width is not None else None)
            except Exception:
                column_widths.append(None)

        default_row_height = None
        try:
            default_row_height = ws.sheet_format.defaultRowHeight
        except Exception:
            default_row_height = None
        for row_index in range(min_row, max_row + 1):
            height = None
            try:
                height = ws.row_dimensions[row_index].height
            except Exception:
                height = None
            if height is None:
                height = default_row_height
            try:
                row_heights.append(float(height) if height is not None else None)
            except Exception:
                row_heights.append(None)
    except Exception:
        return None, None
    return column_widths, row_heights


def _has_any_border(borders):
    try:
        for row in borders or []:
            for cell in row or []:
                if cell.get("top") or cell.get("right") or cell.get("bottom") or cell.get("left"):
                    return True
    except Exception:
        pass
    return False


def _add_merged_range(merged_ranges, min_col, min_row, max_col, max_row, selected_min_col, selected_min_row, selected_max_col, selected_max_row):
    try:
        if max_col < selected_min_col or min_col > selected_max_col:
            return
        if max_row < selected_min_row or min_row > selected_max_row:
            return

        clipped_min_col = max(min_col, selected_min_col)
        clipped_max_col = min(max_col, selected_max_col)
        clipped_min_row = max(min_row, selected_min_row)
        clipped_max_row = min(max_row, selected_max_row)
        if clipped_max_col <= clipped_min_col and clipped_max_row <= clipped_min_row:
            return

        merged_ranges.append({
            "min_col": int(clipped_min_col - selected_min_col),
            "min_row": int(clipped_min_row - selected_min_row),
            "max_col": int(clipped_max_col - selected_min_col),
            "max_row": int(clipped_max_row - selected_min_row),
        })
    except Exception:
        pass


def _read_cell_values_and_borders_via_openpyxl(file_path, worksheet_name, address):
    wb = None
    try:
        import openpyxl
        try:
            from openpyxl.utils.cell import range_boundaries
        except Exception:
            range_boundaries = None

        wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
        if worksheet_name not in wb.sheetnames:
            wb.close()
            return None, "Worksheet was not found by openpyxl."
        ws = wb[worksheet_name]

        if range_boundaries:
            min_col, min_row, max_col, max_row = range_boundaries(address)
        else:
            bounds = _address_to_bounds(address)
            if not bounds:
                wb.close()
                return None, "Openpyxl could not parse address '%s'." % safe_unicode(address)
            min_col, min_row, max_col, max_row = bounds

        data = []
        borders = []
        cell_styles = []
        font_counts = {}
        for row_index in range(min_row, max_row + 1):
            row_values = []
            row_borders = []
            row_styles = []
            for col_index in range(min_col, max_col + 1):
                cell = ws.cell(row=row_index, column=col_index)
                row_values.append(_clean_cell_value(cell.value))
                row_borders.append(_cell_border_dict(cell))
                row_styles.append(_cell_style_dict(cell))
                try:
                    _add_font_count(font_counts, cell.font.name)
                except Exception:
                    pass
            data.append(row_values)
            borders.append(row_borders)
            cell_styles.append(row_styles)
        merged_ranges = []
        try:
            for merged_range in ws.merged_cells.ranges:
                try:
                    try:
                        merged_min_col = int(merged_range.min_col)
                        merged_min_row = int(merged_range.min_row)
                        merged_max_col = int(merged_range.max_col)
                        merged_max_row = int(merged_range.max_row)
                    except Exception:
                        bounds = _address_to_bounds(safe_unicode(merged_range))
                        if not bounds:
                            continue
                        merged_min_col, merged_min_row, merged_max_col, merged_max_row = bounds
                    _add_merged_range(
                        merged_ranges,
                        merged_min_col,
                        merged_min_row,
                        merged_max_col,
                        merged_max_row,
                        min_col,
                        min_row,
                        max_col,
                        max_row,
                    )
                except Exception:
                    pass
        except Exception:
            pass
        column_widths, row_heights = _read_openpyxl_dimensions(ws, min_col, min_row, max_col, max_row)
        dominant_region_font = _dominant_font_from_counts(font_counts)
        wb.close()
        return ExcelTableData(data, borders, True, u"openpyxl", merged_ranges, True, column_widths, row_heights, column_widths is not None or row_heights is not None, u"openpyxl", dominant_region_font, cell_styles if _has_useful_cell_styles(cell_styles) else None), ""
    except Exception as ex:
        try:
            if wb:
                wb.close()
        except Exception:
            pass
        return None, safe_unicode(ex)


def _read_shared_strings(zip_file):
    try:
        import xml.etree.ElementTree as ET
        if 'xl/sharedStrings.xml' not in zip_file.namelist():
            return []
        with zip_file.open('xl/sharedStrings.xml') as f:
            root = ET.parse(f).getroot()
        strings = []
        for si in root:
            parts = []
            for elem in si.iter():
                if elem.tag.endswith('t') and elem.text is not None:
                    parts.append(safe_unicode(elem.text))
            strings.append(u"".join(parts))
        return strings
    except Exception:
        return []


def _get_cell_text_from_xml(cell, shared_strings):
    try:
        cell_type = cell.get('t')
        if cell_type == 'inlineStr':
            parts = []
            for elem in cell.iter():
                if elem.tag.endswith('t') and elem.text is not None:
                    parts.append(safe_unicode(elem.text))
            return _clean_cell_value(u"".join(parts))

        value_elem = None
        for child in cell:
            if child.tag.endswith('v'):
                value_elem = child
                break
        if value_elem is None or value_elem.text is None:
            return u""

        raw = safe_unicode(value_elem.text)
        if cell_type == 's':
            try:
                index = int(raw)
                if 0 <= index < len(shared_strings):
                    return _clean_cell_value(shared_strings[index])
            except Exception:
                return u""
        return _clean_cell_value(raw)
    except Exception:
        return u""


def _read_border_styles_from_zip(zip_file):
    try:
        import xml.etree.ElementTree as ET
        if 'xl/styles.xml' not in zip_file.namelist():
            return None
        with zip_file.open('xl/styles.xml') as f:
            root = ET.parse(f).getroot()

        borders = []
        border_parent = None
        cell_xfs = []
        for elem in root:
            if elem.tag.endswith('borders'):
                border_parent = elem
            elif elem.tag.endswith('cellXfs'):
                cell_xfs = list(elem)

        if border_parent is None:
            return None

        for border in list(border_parent):
            item = {"top": False, "right": False, "bottom": False, "left": False}
            for side in list(border):
                name = side.tag.split('}')[-1]
                if name in item and side.get('style'):
                    item[name] = True
            borders.append(item)

        style_borders = {}
        for index, xf in enumerate(cell_xfs):
            try:
                border_id = int(xf.get('borderId') or 0)
                if 0 <= border_id < len(borders):
                    style_borders[index] = borders[border_id]
            except Exception:
                pass
        return style_borders
    except Exception:
        return None


def _get_cell_border_from_xml(cell, style_borders):
    try:
        if not style_borders:
            return {"top": False, "right": False, "bottom": False, "left": False}
        style_index = int(cell.get('s') or 0)
        if style_index in style_borders:
            item = style_borders[style_index]
            return {
                "top": bool(item.get("top")),
                "right": bool(item.get("right")),
                "bottom": bool(item.get("bottom")),
                "left": bool(item.get("left")),
            }
    except Exception:
        pass
    return {"top": False, "right": False, "bottom": False, "left": False}


def _read_merged_ranges_from_sheet_xml(root, selected_min_col, selected_min_row, selected_max_col, selected_max_row):
    merged_ranges = []
    try:
        for elem in root.iter():
            if not elem.tag.endswith('mergeCell'):
                continue
            ref = elem.get('ref')
            if not ref:
                continue
            bounds = _address_to_bounds(ref)
            if not bounds:
                continue
            min_col, min_row, max_col, max_row = bounds
            _add_merged_range(
                merged_ranges,
                min_col,
                min_row,
                max_col,
                max_row,
                selected_min_col,
                selected_min_row,
                selected_max_col,
                selected_max_row,
            )
    except Exception:
        pass
    return merged_ranges


def _read_dimensions_from_sheet_xml(root, min_col, min_row, max_col, max_row):
    column_widths = []
    row_heights = []
    default_col_width = None
    default_row_height = None
    try:
        for elem in root.iter():
            if elem.tag.endswith('sheetFormatPr'):
                try:
                    if elem.get('defaultColWidth') is not None:
                        default_col_width = float(elem.get('defaultColWidth'))
                except Exception:
                    default_col_width = None
                try:
                    if elem.get('defaultRowHeight') is not None:
                        default_row_height = float(elem.get('defaultRowHeight'))
                except Exception:
                    default_row_height = None
                break

        widths_by_col = {}
        for elem in root.iter():
            if not elem.tag.endswith('col'):
                continue
            try:
                col_min = int(elem.get('min'))
                col_max = int(elem.get('max'))
                width = elem.get('width')
                width = float(width) if width is not None else default_col_width
                for col_index in range(col_min, col_max + 1):
                    widths_by_col[col_index] = width
            except Exception:
                pass

        heights_by_row = {}
        for elem in root.iter():
            if not elem.tag.endswith('row'):
                continue
            try:
                row_index = int(elem.get('r'))
                height = elem.get('ht')
                if height is not None:
                    heights_by_row[row_index] = float(height)
            except Exception:
                pass

        for col_index in range(min_col, max_col + 1):
            column_widths.append(widths_by_col.get(col_index, default_col_width))
        for row_index in range(min_row, max_row + 1):
            row_heights.append(heights_by_row.get(row_index, default_row_height))
    except Exception:
        return None, None
    return column_widths, row_heights


def _read_cell_values_via_zipfile(file_path, worksheet_name, address):
    try:
        import zipfile
        import xml.etree.ElementTree as ET

        bounds = _address_to_bounds(address)
        if not bounds:
            return None, "ZIP reader could not parse address '%s'." % safe_unicode(address)
        min_col, min_row, max_col, max_row = bounds

        if not zipfile.is_zipfile(file_path):
            return None, "File is not an XLSX/XLSM ZIP package."

        with zipfile.ZipFile(file_path, 'r') as z:
            sheet_path = _get_sheet_path_from_workbook_relationships(z, worksheet_name)
            if not sheet_path or sheet_path not in z.namelist():
                return None, "Worksheet XML was not found in the workbook."
            shared_strings = _read_shared_strings(z)
            with z.open(sheet_path) as f:
                root = ET.parse(f).getroot()

        values = {}
        for cell in root.iter():
            ref = cell.get('r')
            if not ref:
                continue
            match = re.match(r'^([A-Z]+)([0-9]+)$', ref.upper())
            if not match:
                continue
            col = _column_letter_to_number(match.group(1))
            row = int(match.group(2))
            if min_col <= col <= max_col and min_row <= row <= max_row:
                values[(row, col)] = _get_cell_text_from_xml(cell, shared_strings)

        data = []
        for row_index in range(min_row, max_row + 1):
            row_values = []
            for col_index in range(min_col, max_col + 1):
                row_values.append(values.get((row_index, col_index), u""))
            data.append(row_values)
        column_widths, row_heights = _read_dimensions_from_sheet_xml(root, min_col, min_row, max_col, max_row)
        return ExcelTableData(data, None, False, u"zip values", None, False, column_widths, row_heights, column_widths is not None or row_heights is not None, u"zip"), ""
    except Exception as ex:
        return None, safe_unicode(ex)


def _read_cell_values_and_borders_via_zipfile(file_path, worksheet_name, address):
    try:
        import zipfile
        import xml.etree.ElementTree as ET

        bounds = _address_to_bounds(address)
        if not bounds:
            return None, "ZIP reader could not parse address '%s'." % safe_unicode(address)
        min_col, min_row, max_col, max_row = bounds

        if not zipfile.is_zipfile(file_path):
            return None, "File is not an XLSX/XLSM ZIP package."

        with zipfile.ZipFile(file_path, 'r') as z:
            sheet_path = _get_sheet_path_from_workbook_relationships(z, worksheet_name)
            if not sheet_path or sheet_path not in z.namelist():
                return None, "Worksheet XML was not found in the workbook."
            shared_strings = _read_shared_strings(z)
            style_borders = _read_border_styles_from_zip(z)
            with z.open(sheet_path) as f:
                root = ET.parse(f).getroot()

        values = {}
        borders = {}
        for cell in root.iter():
            ref = cell.get('r')
            if not ref:
                continue
            match = re.match(r'^([A-Z]+)([0-9]+)$', ref.upper())
            if not match:
                continue
            col = _column_letter_to_number(match.group(1))
            row = int(match.group(2))
            if min_col <= col <= max_col and min_row <= row <= max_row:
                values[(row, col)] = _get_cell_text_from_xml(cell, shared_strings)
                borders[(row, col)] = _get_cell_border_from_xml(cell, style_borders)

        data = []
        border_rows = []
        empty_border = {"top": False, "right": False, "bottom": False, "left": False}
        for row_index in range(min_row, max_row + 1):
            row_values = []
            row_borders = []
            for col_index in range(min_col, max_col + 1):
                row_values.append(values.get((row_index, col_index), u""))
                row_borders.append(borders.get((row_index, col_index), empty_border))
            data.append(row_values)
            border_rows.append(row_borders)
        merged_ranges = _read_merged_ranges_from_sheet_xml(root, min_col, min_row, max_col, max_row)
        column_widths, row_heights = _read_dimensions_from_sheet_xml(root, min_col, min_row, max_col, max_row)
        return ExcelTableData(data, border_rows, style_borders is not None, u"zip", merged_ranges, True, column_widths, row_heights, column_widths is not None or row_heights is not None, u"zip"), ""
    except Exception as ex:
        return None, safe_unicode(ex)


def _read_cell_values_via_com(file_path, worksheet_name, address):
    excel = None
    workbook = None
    sheet = None
    cell_range = None
    try:
        excel_type = System.Type.GetTypeFromProgID("Excel.Application")
        if excel_type is None:
            return None, "Excel COM is not available."
        excel = System.Activator.CreateInstance(excel_type)
        _set_excel_quiet(excel)
        workbook = excel.Workbooks.Open(file_path, False, True)
        sheet = _find_worksheet_by_name(workbook, worksheet_name)
        if sheet is None:
            return None, "Worksheet was not found by Excel COM."
        if safe_unicode(address) == USED_RANGE_KEY:
            cell_range = sheet.UsedRange
        else:
            cell_range = sheet.Range(address)
        values = cell_range.Value2
        rows = int(cell_range.Rows.Count)
        cols = int(cell_range.Columns.Count)
        data = []
        for r in range(1, rows + 1):
            row_values = []
            for c in range(1, cols + 1):
                try:
                    if rows == 1 and cols == 1:
                        value = values
                    else:
                        value = values[r, c]
                except Exception:
                    try:
                        value = cell_range.Cells(r, c).Value2
                    except Exception:
                        value = None
                row_values.append(_clean_cell_value(value))
            data.append(row_values)
        return data, ""
    except Exception as ex:
        return None, safe_unicode(ex)
    finally:
        _release_com_object(cell_range)
        _release_com_object(sheet)
        try:
            if workbook:
                workbook.Close(False)
        except Exception:
            pass
        _release_com_object(workbook)
        try:
            if excel:
                excel.Quit()
        except Exception:
            pass
        _release_com_object(excel)


def read_excel_table_values(file_path, worksheet_name, region):
    """
    Return (data, row_count, column_count) for the requested worksheet region.
    Data is a 2D list of Unicode strings.
    """
    original_file_path = file_path
    if not original_file_path or not os.path.exists(original_file_path):
        return [], 0, 0

    worksheet_name = clean_hidden_unicode(safe_unicode(worksheet_name))
    region = clean_hidden_unicode(safe_unicode(region))
    address, resolve_method, resolve_reason = _resolve_region_address_with_method(original_file_path, worksheet_name, region)

    data = None
    attempted = []
    if address:
        data, reason = _read_cell_values_and_borders_via_openpyxl(original_file_path, worksheet_name, address)
        attempted.append(("openpyxl borders", reason))
        if data is None:
            ext = os.path.splitext(safe_unicode(original_file_path))[1].lower()
            if ext in ('.xlsx', '.xlsm'):
                data, reason = _read_cell_values_and_borders_via_zipfile(original_file_path, worksheet_name, address)
                attempted.append(("zip borders", reason))
        if data is None:
            data, reason = _read_cell_values_via_openpyxl(original_file_path, worksheet_name, address)
            attempted.append(("openpyxl values", reason))
            if data is not None and not isinstance(data, ExcelTableData):
                data = ExcelTableData(data, None, False, u"openpyxl values")
        if data is None:
            ext = os.path.splitext(safe_unicode(original_file_path))[1].lower()
            if ext in ('.xlsx', '.xlsm'):
                data, reason = _read_cell_values_via_zipfile(original_file_path, worksheet_name, address)
                attempted.append(("zip values", reason))
                if data is not None and not isinstance(data, ExcelTableData):
                    data = ExcelTableData(data, None, False, u"zip values")
        if data is None:
            data, reason = _read_cell_values_via_com(original_file_path, worksheet_name, address)
            attempted.append(("COM", reason))
            if data is not None:
                data = ExcelTableData(data, None, False, u"COM")
    if data is None:
        data = ExcelTableData([], None, False, u"")

    row_count = len(data)
    column_count = 0
    for row in data:
        try:
            column_count = max(column_count, len(row))
        except Exception:
            pass
    for row in data:
        while len(row) < column_count:
            row.append(u"")
    if row_count <= 0 or column_count <= 0:
        print("Table Importer Excel read debug:")
        print("  file path: %s" % safe_unicode(original_file_path))
        print("  worksheet: %s" % safe_unicode(worksheet_name))
        print("  region display: %s" % safe_unicode(region))
        print("  resolved address: %s" % safe_unicode(address))
        print("  method used: resolve=%s" % safe_unicode(resolve_method))
        if attempted:
            for method, reason in attempted:
                print("  reader attempted: %s | reason: %s" % (safe_unicode(method), safe_unicode(reason)))
        else:
            print("  reader attempted: none")
        reason = resolve_reason or "No cells returned."
        print("  reason: %s" % safe_unicode(reason))
    return data, row_count, column_count


def read_excel_region_data(file_path, worksheet_name, region_display):
    data, row_count, column_count = read_excel_table_values(file_path, worksheet_name, region_display)
    return data

